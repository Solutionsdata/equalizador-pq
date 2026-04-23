"""
Serviço de geração e importação de arquivos Excel para PQ e Propostas.

Layouts:
  Planilha PQ  (modelo_pq)      — 12 colunas em azul  (somente leitura para proponentes)
  Proposta      (modelo_proposta)— 12 colunas azuis + 4 laranja (Sem REIDI) + 4 verdes (Com REIDI)
"""

import io
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.pq_item import PQItem
from app.models.proposal_item import ProposalItem

# ── Paleta de cores ───────────────────────────────────────────────────────────
AZUL_TITULO   = "1E3A5F"
AZUL_HEADER   = "2563EB"
AZUL_LINHA_A  = "DBEAFE"
AZUL_LINHA_B  = "EFF6FF"
VERDE_HEADER  = "15803D"
VERDE_LINHA_A = "DCFCE7"
VERDE_LINHA_B = "F0FDF4"
LARANJA_HEADER  = "C2410C"
LARANJA_LINHA_A = "FFEDD5"
LARANJA_LINHA_B = "FFF7ED"
AMARELO_AVISO = "FEF9C3"
CINZA_INFO    = "F3F4F6"
BRANCO        = "FFFFFF"

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _font(color="000000", bold=False, size=9) -> Font:
    return Font(color=color, bold=bold, size=size, name="Calibri")


def _border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Definição das colunas ─────────────────────────────────────────────────────
# (rótulo, campo_model, largura, formato)
# campo "_total_rf" = coluna calculada (qtd × preco_referencia)
PQ_COLS = [
    ("Item",             "numero_item",        8,  "text"),
    ("Localidade",       "localidade",        20,  "text"),
    ("Disciplina",       "disciplina",        16,  "text"),
    ("Categoria",        "categoria",         20,  "text"),
    ("Código",           "codigo",            13,  "text"),
    ("Descrição",        "descricao",         52,  "text"),
    ("Unidade Medida",   "unidade",           14,  "text"),
    ("Quantidade",       "quantidade",        13,  "num4"),
    ("Referência",       "referencia_codigo", 13,  "text"),
    ("Preço Unit. RF",   "preco_referencia",  16,  "brl4"),
    ("Preço Total RF",   "_total_rf",         16,  "brl2"),
    ("Observação",       "observacao",        30,  "text"),
]

# Posições fixas das colunas-chave na PQ (1-based)
PQ_QTD_COL    = 8   # Quantidade
PQ_PUNIT_COL  = 10  # Preço Unit. RF
PQ_TOTAL_COL  = 11  # Preço Total RF (fórmula)
PQ_NCOLS      = len(PQ_COLS)  # 12

# Proposta: SEM REIDI começa em col 13, COM REIDI em col 17
SEM_START = PQ_NCOLS + 1   # 13
COM_START = PQ_NCOLS + 5   # 17
TOTAL_PROP_COLS = PQ_NCOLS + 8  # 20

FMT = {
    "text":  "@",
    "num4":  "#,##0.0000",
    "brl4":  'R$ #,##0.0000',
    "brl2":  'R$ #,##0.00',
    "pct":   '0.00"%"',
}


# ── Geração: Modelo PQ ────────────────────────────────────────────────────────

def gerar_modelo_pq(project_name: str, pq_items: Optional[List[PQItem]] = None) -> io.BytesIO:
    """
    Gera o Excel da Planilha de Quantitativos (12 colunas).
    Se `pq_items` for fornecido, exporta os dados; caso contrário exporta template vazio.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha PQ"
    ws.sheet_view.showGridLines = False

    last_col = get_column_letter(PQ_NCOLS)

    # Linha 1 — Título
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"PLANILHA DE QUANTIDADES  ·  {project_name.upper()}"
    c.fill = _fill(AZUL_TITULO)
    c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Linha 2 — Aviso
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (
        "MODELO OFICIAL  ·  Não altere os cabeçalhos nem a ordem das colunas  ·  "
        "Preencha a partir da linha 4"
    )
    c.fill = _fill(AMARELO_AVISO)
    c.font = _font("78350F", size=9)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 16

    # Linha 3 — Cabeçalhos
    for ci, (label, _, width, _) in enumerate(PQ_COLS, start=1):
        c = ws.cell(row=3, column=ci, value=label)
        c.fill = _fill(AZUL_HEADER)
        c.font = _font("FFFFFF", bold=True, size=9)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    qtd_letter  = get_column_letter(PQ_QTD_COL)
    punit_letter = get_column_letter(PQ_PUNIT_COL)

    # Linhas de dados
    rows = pq_items if pq_items else []
    for ri, item in enumerate(rows, start=4):
        alt = ri % 2 == 0
        for ci, (_, field, _, fmt) in enumerate(PQ_COLS, start=1):
            if field == "_total_rf":
                val = f"={qtd_letter}{ri}*{punit_letter}{ri}"
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = FMT["brl2"]
                c.alignment = _align("right")
            else:
                raw = getattr(item, field, None)
                val = float(raw) if raw is not None and fmt in ("num4", "brl4") else raw
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = FMT.get(fmt, "@")
                c.alignment = _align("right" if fmt != "text" else "left")
            c.fill = _fill(AZUL_LINHA_A if alt else AZUL_LINHA_B)
            c.font = _font(size=9)
            c.border = _border()

    # Template vazio: 50 linhas com fórmulas
    if not rows:
        for ri in range(4, 54):
            alt = ri % 2 == 0
            for ci, (_, field, _, fmt) in enumerate(PQ_COLS, start=1):
                if field == "_total_rf":
                    c = ws.cell(row=ri, column=ci,
                                value=f"={qtd_letter}{ri}*{punit_letter}{ri}")
                    c.number_format = FMT["brl2"]
                    c.alignment = _align("right")
                else:
                    c = ws.cell(row=ri, column=ci)
                    c.number_format = FMT.get(fmt, "@")
                c.fill = _fill(AZUL_LINHA_A if alt else AZUL_LINHA_B)
                c.border = _border()

    return _save(wb)


# ── Geração: Modelo Proposta ──────────────────────────────────────────────────

def gerar_modelo_proposta(
    project_name: str,
    pq_items: List[PQItem],
    empresa: str = "",
    bdi_global: float = 0.0,
    proposal_items: Optional[List[ProposalItem]] = None,
) -> io.BytesIO:
    """
    Gera o Excel de Proposta Comercial.
    12 colunas PQ (azul) + 4 colunas Sem REIDI (laranja) + 4 colunas Com REIDI (verde).
    Analytics usa o Custo Total BDI COM REIDI.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposta"
    ws.sheet_view.showGridLines = False

    last_col = get_column_letter(TOTAL_PROP_COLS)
    pq_end_col = get_column_letter(PQ_NCOLS)

    # Col letters para as colunas de preço
    qtd_letter   = get_column_letter(PQ_QTD_COL)
    punit_letter = get_column_letter(PQ_PUNIT_COL)
    col_cud_sem       = SEM_START      # 13
    col_bdi_sem       = SEM_START + 1  # 14
    col_cud_bdi_sem   = SEM_START + 2  # 15
    col_total_sem     = SEM_START + 3  # 16
    col_cud_com       = COM_START      # 17
    col_bdi_com       = COM_START + 1  # 18
    col_cud_bdi_com   = COM_START + 2  # 19
    col_total_com     = COM_START + 3  # 20

    # Linha 1 — Título
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"PROPOSTA COMERCIAL  ·  {project_name.upper()}"
    c.fill = _fill(AZUL_TITULO)
    c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Linha 2 — Identificação
    half = PQ_NCOLS // 2
    ws.merge_cells(f"A2:{get_column_letter(half)}2")
    c = ws["A2"]
    c.value = f"Empresa / Proponente:  {empresa or '___________________________________'}"
    c.fill = _fill(CINZA_INFO)
    c.font = _font(bold=True, size=10)
    c.alignment = _align("left")

    ws.merge_cells(f"{get_column_letter(half+1)}2:{last_col}2")
    c = ws.cell(row=2, column=half + 1)
    c.value = f"BDI Global aplicado (%):  {bdi_global:.2f}"
    c.fill = _fill(CINZA_INFO)
    c.font = _font(bold=True, size=10)
    c.alignment = _align("right")
    ws.row_dimensions[2].height = 22

    # Linha 3 — Instrução
    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]
    c.value = (
        "INSTRUÇÃO:  Preencha SOMENTE as colunas em LARANJA (Sem REIDI) e VERDE (Com REIDI).  "
        "Não altere as colunas em AZUL.  Salve e devolva este arquivo."
    )
    c.fill = _fill(AMARELO_AVISO)
    c.font = _font("78350F", size=9)
    c.alignment = _align("center")
    ws.row_dimensions[3].height = 16

    # Linha 4 — Cabeçalhos de grupo
    ws.merge_cells(f"A4:{pq_end_col}4")
    c = ws["A4"]
    c.value = "PLANILHA DE QUANTITATIVOS"
    c.fill = _fill(AZUL_TITULO)
    c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("center")
    c.border = _border()

    ws.merge_cells(f"{get_column_letter(col_cud_sem)}4:{get_column_letter(col_total_sem)}4")
    c = ws.cell(row=4, column=col_cud_sem, value="SEM REIDI")
    c.fill = _fill(LARANJA_HEADER)
    c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("center")
    c.border = _border()

    ws.merge_cells(f"{get_column_letter(col_cud_com)}4:{get_column_letter(col_total_com)}4")
    c = ws.cell(row=4, column=col_cud_com, value="COM REIDI")
    c.fill = _fill(VERDE_HEADER)
    c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("center")
    c.border = _border()
    ws.row_dimensions[4].height = 20

    # Linha 5 — Cabeçalhos individuais
    for ci, (label, _, width, _) in enumerate(PQ_COLS, start=1):
        c = ws.cell(row=5, column=ci, value=label)
        c.fill = _fill(AZUL_HEADER)
        c.font = _font("FFFFFF", bold=True, size=9)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    sem_headers = [
        ("Custo Unit. Direto", 18),
        ("BDI (%)",            10),
        ("Custo Unit. c/BDI",  18),
        ("Custo Total c/BDI",  20),
    ]
    for i, (label, width) in enumerate(sem_headers):
        ci = col_cud_sem + i
        c = ws.cell(row=5, column=ci, value=label)
        c.fill = _fill(LARANJA_HEADER)
        c.font = _font("FFFFFF", bold=True, size=9)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    com_headers = [
        ("Custo Unit. Direto", 18),
        ("BDI (%)",            10),
        ("Custo Unit. c/BDI",  18),
        ("Custo Total c/BDI",  20),
    ]
    for i, (label, width) in enumerate(com_headers):
        ci = col_cud_com + i
        c = ws.cell(row=5, column=ci, value=label)
        c.fill = _fill(VERDE_HEADER)
        c.font = _font("FFFFFF", bold=True, size=9)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[5].height = 20
    ws.freeze_panes = "A6"

    # Mapa de preços existentes
    price_map: dict[int, ProposalItem] = {}
    if proposal_items:
        for pi in proposal_items:
            price_map[pi.pq_item_id] = pi

    data_start = 6

    # Linhas de dados
    for ri, item in enumerate(pq_items, start=data_start):
        alt = ri % 2 == 0
        pi = price_map.get(item.id)

        # ── Colunas PQ (azul) ────────────────────────────────────────────────
        for ci, (_, field, _, fmt) in enumerate(PQ_COLS, start=1):
            if field == "_total_rf":
                val = f"={qtd_letter}{ri}*{punit_letter}{ri}"
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = FMT["brl2"]
                c.alignment = _align("right")
            else:
                raw = getattr(item, field, None)
                val = float(raw) if raw is not None and fmt in ("num4", "brl4") else raw
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = FMT.get(fmt, "@")
                c.alignment = _align("right" if fmt != "text" else "left")
            c.fill = _fill(AZUL_LINHA_A if alt else AZUL_LINHA_B)
            c.font = _font(size=9)
            c.border = _border()

        # ── SEM REIDI (laranja) ──────────────────────────────────────────────
        cud_sem_val = float(pi.preco_unitario) if pi and pi.preco_unitario else None
        bdi_sem_val = float(pi.bdi) if pi and pi.bdi else None

        cud_sem_ref = f"{get_column_letter(col_cud_sem)}{ri}"
        bdi_sem_ref = f"{get_column_letter(col_bdi_sem)}{ri}"

        c = ws.cell(row=ri, column=col_cud_sem, value=cud_sem_val)
        c.fill = _fill(LARANJA_LINHA_A if alt else LARANJA_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["brl4"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_bdi_sem, value=bdi_sem_val)
        c.fill = _fill(LARANJA_LINHA_A if alt else LARANJA_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["pct"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_cud_bdi_sem,
                    value=f"={cud_sem_ref}*(1+{bdi_sem_ref}/100)")
        c.fill = _fill(LARANJA_LINHA_A if alt else LARANJA_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["brl4"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_total_sem,
                    value=f"={qtd_letter}{ri}*{cud_sem_ref}*(1+{bdi_sem_ref}/100)")
        c.fill = _fill(LARANJA_LINHA_A if alt else LARANJA_LINHA_B)
        c.font = _font("C2410C", bold=True, size=9); c.border = _border()
        c.number_format = FMT["brl2"]; c.alignment = _align("right")

        # ── COM REIDI (verde) ────────────────────────────────────────────────
        cud_com_val = float(pi.custo_unit_com_reidi) if pi and pi.custo_unit_com_reidi else None
        bdi_com_val = float(pi.bdi_com_reidi) if pi and pi.bdi_com_reidi else None

        cud_com_ref = f"{get_column_letter(col_cud_com)}{ri}"
        bdi_com_ref = f"{get_column_letter(col_bdi_com)}{ri}"

        c = ws.cell(row=ri, column=col_cud_com, value=cud_com_val)
        c.fill = _fill(VERDE_LINHA_A if alt else VERDE_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["brl4"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_bdi_com, value=bdi_com_val)
        c.fill = _fill(VERDE_LINHA_A if alt else VERDE_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["pct"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_cud_bdi_com,
                    value=f"={cud_com_ref}*(1+{bdi_com_ref}/100)")
        c.fill = _fill(VERDE_LINHA_A if alt else VERDE_LINHA_B)
        c.font = _font(size=9); c.border = _border()
        c.number_format = FMT["brl4"]; c.alignment = _align("right")

        c = ws.cell(row=ri, column=col_total_com,
                    value=f"={qtd_letter}{ri}*{cud_com_ref}*(1+{bdi_com_ref}/100)")
        c.fill = _fill(VERDE_LINHA_A if alt else VERDE_LINHA_B)
        c.font = _font("15803D", bold=True, size=9); c.border = _border()
        c.number_format = FMT["brl2"]; c.alignment = _align("right")

    # ── Linha de total ────────────────────────────────────────────────────────
    total_row = len(pq_items) + data_start
    ws.merge_cells(f"A{total_row}:{pq_end_col}{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "TOTAL GERAL DA PROPOSTA"
    c.fill = _fill(AZUL_TITULO)
    c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("right")
    c.border = _border()

    # Total SEM REIDI
    for ci in (col_cud_sem, col_bdi_sem, col_cud_bdi_sem):
        c = ws.cell(row=total_row, column=ci)
        c.fill = _fill(LARANJA_HEADER); c.border = _border()

    tot_sem_ltr = get_column_letter(col_total_sem)
    c = ws.cell(row=total_row, column=col_total_sem,
                value=f"=SUM({tot_sem_ltr}{data_start}:{tot_sem_ltr}{total_row-1})")
    c.fill = _fill(LARANJA_HEADER)
    c.font = _font("FFFFFF", bold=True, size=11)
    c.number_format = FMT["brl2"]
    c.alignment = _align("right"); c.border = _border()

    # Total COM REIDI
    for ci in (col_cud_com, col_bdi_com, col_cud_bdi_com):
        c = ws.cell(row=total_row, column=ci)
        c.fill = _fill(AZUL_TITULO); c.border = _border()

    tot_com_ltr = get_column_letter(col_total_com)
    c = ws.cell(row=total_row, column=col_total_com,
                value=f"=SUM({tot_com_ltr}{data_start}:{tot_com_ltr}{total_row-1})")
    c.fill = _fill(AZUL_TITULO)
    c.font = _font("FFFFFF", bold=True, size=11)
    c.number_format = FMT["brl2"]
    c.alignment = _align("right"); c.border = _border()

    ws.row_dimensions[total_row].height = 22

    return _save(wb)


# ── Importação: PQ ────────────────────────────────────────────────────────────

def importar_pq_excel(file_bytes: bytes) -> List[dict]:
    """
    Lê um Excel (modelo PQ ou qualquer planilha com cabeçalhos compatíveis)
    e retorna lista de dicts prontos para inserção no banco.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    MAPA_HEADERS = {
        "item": "numero_item", "nº item": "numero_item", "número item": "numero_item",
        "localidade": "localidade",
        "disciplina": "disciplina",
        "categoria": "categoria",
        "código": "codigo", "codigo": "codigo",
        "descrição": "descricao", "descricao": "descricao",
        "unidade": "unidade", "un": "unidade", "unidade medida": "unidade",
        "quantidade": "quantidade", "qtd": "quantidade",
        "referência": "referencia_codigo", "referencia": "referencia_codigo",
        "ref.": "referencia_codigo",
        "preço unit. rf": "preco_referencia", "preço unitário rf": "preco_referencia",
        "preço ref. (r$)": "preco_referencia", "preço ref.": "preco_referencia",
        "preco referencia": "preco_referencia", "preço referência": "preco_referencia",
        "observação": "observacao", "observacao": "observacao", "obs.": "observacao",
        # Ignora a coluna de fórmula "Preço Total RF"
    }

    header_row, col_map = _find_header(ws, MAPA_HEADERS, required={"numero_item", "descricao"})

    items = []
    for row in ws.iter_rows(min_row=header_row + 1):
        rd = {col_map[c.column]: c.value for c in row if c.column in col_map}
        num = str(rd.get("numero_item") or "").strip()
        desc = str(rd.get("descricao") or "").strip()
        if not num or not desc:
            continue

        items.append({
            "numero_item":       num,
            "localidade":        _str(rd.get("localidade")),
            "disciplina":        _str(rd.get("disciplina")),
            "categoria":         _str(rd.get("categoria")),
            "codigo":            _str(rd.get("codigo")),
            "descricao":         desc,
            "unidade":           _str(rd.get("unidade")) or "un",
            "quantidade":        _float(rd.get("quantidade")) or 0,
            "referencia_codigo": _str(rd.get("referencia_codigo")),
            "preco_referencia":  _float(rd.get("preco_referencia")),
            "observacao":        _str(rd.get("observacao")),
            "ordem":             len(items),
        })

    if not items:
        raise ValueError("Nenhum item válido encontrado. Verifique se o arquivo usa o modelo padrão.")
    return items


# ── Importação: Proposta ──────────────────────────────────────────────────────

def importar_proposta_excel(file_bytes: bytes, pq_items: List[PQItem]) -> List[dict]:
    """
    Lê um Excel de proposta preenchido pelo proponente.
    Layout fixo: 12 cols PQ + 4 cols Sem REIDI (13-16) + 4 cols Com REIDI (17-20).
    Editable: col 13 (CUD sem), 14 (BDI sem), 17 (CUD com), 18 (BDI com).
    Retorna lista de {pq_item_id, preco_unitario, bdi, custo_unit_com_reidi, bdi_com_reidi, ...}.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    pq_map = {str(item.numero_item).strip(): item.id for item in pq_items}

    # Encontra a linha de cabeçalho individual (onde col A tem "Item")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        first = str(row[0].value or "").strip().lower()
        if first in ("item", "nº item", "número item"):
            header_row = row[0].row
            break

    if header_row is None:
        # Fallback: tenta qualquer linha com "item" em col 1
        for row in ws.iter_rows(min_row=1, max_row=15):
            for cell in row:
                if cell.column == 1 and str(cell.value or "").strip().lower() == "item":
                    header_row = cell.row
                    break
            if header_row:
                break

    if header_row is None:
        raise ValueError(
            "Cabeçalho 'Item' não encontrado. Verifique se o arquivo usa o modelo de proposta."
        )

    # Posições fixas (1-based) — correspondem ao template gerado
    COL_NUM  = 1   # numero_item
    COL_DESC = 6   # Descrição (scope capture)
    COL_UN   = 7   # Unidade Medida
    COL_QTD  = 8   # Quantidade
    COL_CUD_SEM = 13
    COL_BDI_SEM = 14
    COL_CUD_COM = 17
    COL_BDI_COM = 18

    result = []
    for row in ws.iter_rows(min_row=header_row + 1):
        # row é zero-indexed na tupla, mas cell.column é 1-based
        def _val(col_1based):
            idx = col_1based - 1
            return row[idx].value if idx < len(row) else None

        num = str(_val(COL_NUM) or "").strip()
        if num not in pq_map:
            continue

        cud_com = _float(_val(COL_CUD_COM))
        cud_sem = _float(_val(COL_CUD_SEM))

        # Exige pelo menos o preço COM REIDI
        if cud_com is None and cud_sem is None:
            continue

        result.append({
            "pq_item_id":            pq_map[num],
            "preco_unitario":        cud_sem,
            "bdi":                   _float(_val(COL_BDI_SEM)),
            "custo_unit_com_reidi":  cud_com,
            "bdi_com_reidi":         _float(_val(COL_BDI_COM)),
            "descricao_proposta":    _str(_val(COL_DESC)),
            "unidade_proposta":      _str(_val(COL_UN)),
            "quantidade_proposta":   _float(_val(COL_QTD)),
        })

    if not result:
        raise ValueError(
            "Nenhum preço encontrado. Verifique se o arquivo usa o modelo de proposta "
            "e se as colunas COM REIDI estão preenchidas (cols 17-18)."
        )
    return result


# ── Relatório de Equalização (multi-abas) ────────────────────────────────────

def gerar_relatorio_equalizacao(
    project_name: str,
    eq_data: dict,
    pareto_data: dict,
) -> io.BytesIO:
    """
    Gera relatório Excel completo com 5 abas.
    Todos os preços usam COM REIDI (já retornado pelo AnalyticsService).
    """
    wb = Workbook()
    proposals = eq_data.get("proposals", [])
    items     = eq_data.get("items", [])

    def _th(ws, row, col, value, bg=AZUL_HEADER, fc="FFFFFF", bold=True, h="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = _fill(bg)
        c.font = _font(fc, bold=bold, size=9)
        c.alignment = _align(h)
        c.border = _border()
        return c

    def _td(ws, row, col, value, bg=BRANCO, fmt="@", h="left", bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = _fill(bg)
        c.font = _font(bold=bold, size=9)
        c.alignment = _align(h)
        c.border = _border()
        c.number_format = fmt
        return c

    def _title_row(ws, ncols, title):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value = title
        c.fill = _fill(AZUL_TITULO)
        c.font = _font("FFFFFF", bold=True, size=12)
        c.alignment = _align("center")
        ws.row_dimensions[1].height = 26

    BRL  = 'R$ #,##0.00'
    PCT  = '0.00"%"'
    NUM  = '#,##0.00'
    LINHAS = ["F2F2F2", BRANCO]

    # ── ABA 1: Resumo ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.sheet_view.showGridLines = False

    sorted_props = sorted(proposals, key=lambda p: p.get("valor_total", 0))
    best_total = sorted_props[0]["valor_total"] if sorted_props else 0

    _title_row(ws1, 7, f"RESUMO DA EQUALIZAÇÃO  ·  {project_name.upper()}")
    headers = ["Posição", "Empresa", "BDI Global (%)", "Total COM REIDI (R$)",
               "Δ vs. Menor (R$)", "Δ vs. Menor (%)", "Status"]
    for ci, h in enumerate(headers, 1):
        _th(ws1, 2, ci, h)
        ws1.column_dimensions[get_column_letter(ci)].width = [8,36,14,22,18,14,14][ci-1]
    ws1.row_dimensions[2].height = 18

    for ri, p in enumerate(sorted_props, start=3):
        bg = LINHAS[ri % 2]
        delta_abs = p["valor_total"] - best_total
        delta_pct = (delta_abs / best_total * 100) if best_total else 0
        _td(ws1, ri, 1, ri-2, bg, h="center", bold=(ri==3))
        _td(ws1, ri, 2, p["empresa"], bg)
        _td(ws1, ri, 3, p.get("bdi_global", 0), bg, PCT, "right")
        c = _td(ws1, ri, 4, p["valor_total"], bg, BRL, "right", bold=(ri==3))
        if ri == 3:
            c.font = _font("15803D", bold=True, size=9)
        _td(ws1, ri, 5, delta_abs if ri > 3 else None, bg, BRL, "right")
        _td(ws1, ri, 6, delta_pct if ri > 3 else None, bg, PCT, "right")
        _td(ws1, ri, 7, p.get("status", ""), bg, "@", "center")

    ws1.freeze_panes = "A3"

    cp_total = 0.0
    for item in items:
        precos = {k: v for k, v in item.get("precos", {}).items() if v is not None}
        if precos:
            min_pu = min(precos.values())
            qty = item.get("quantidade", 0)
            cp_total += float(qty) * float(min_pu)

    cr = len(sorted_props) + 4
    ws1.merge_cells(f"A{cr}:C{cr}")
    c = ws1[f"A{cr}"]
    c.value = "POTENCIAL CHERRY PICKING (melhor preço por item)"
    c.fill = _fill("FEF9C3"); c.font = _font("78350F", bold=True, size=9)
    c.alignment = _align("left")
    c = ws1.cell(row=cr, column=4, value=cp_total)
    c.fill = _fill("FEF9C3"); c.font = _font("92400E", bold=True, size=9)
    c.number_format = BRL; c.alignment = _align("right")
    if best_total:
        ec = best_total - cp_total
        ws1.cell(row=cr, column=5, value=ec).number_format = BRL
        ws1.cell(row=cr, column=5).font = _font("15803D", bold=True, size=9)
        ws1.cell(row=cr, column=5).fill = _fill("FEF9C3")
        ws1.cell(row=cr, column=5).alignment = _align("right")
        pct_ec = ws1.cell(row=cr, column=6, value=(ec/best_total*100) if best_total else 0)
        pct_ec.number_format = PCT; pct_ec.font = _font("15803D", bold=True, size=9)
        pct_ec.fill = _fill("FEF9C3"); pct_ec.alignment = _align("right")

    # ── ABA 2: Comparativo ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Comparativo de Preços")
    ws2.sheet_view.showGridLines = False

    fixed_cols = ["Item", "Código", "Descrição", "Un", "Qtd", "Ref. Unit. (R$)", "Ref. Total (R$)"]
    price_cols  = []
    for p in proposals:
        price_cols += [f"{p['empresa'][:20]} — P.Unit. COM REIDI", f"{p['empresa'][:20]} — Total COM REIDI"]
    extra_cols = ["Menor P.Unit.", "Maior P.Unit.", "Preço Médio", "Desvio Padrão", "Categoria", "Disciplina"]
    all_cols = fixed_cols + price_cols + extra_cols

    _title_row(ws2, len(all_cols), f"COMPARATIVO DE PREÇOS (COM REIDI)  ·  {project_name.upper()}")
    for ci, h in enumerate(all_cols, 1):
        is_fixed = ci <= len(fixed_cols)
        _th(ws2, 2, ci, h, bg=AZUL_HEADER if is_fixed else "2D6A4F")
        if ci == 3:
            ws2.column_dimensions[get_column_letter(ci)].width = 48
        elif ci in (5, 6, 7):
            ws2.column_dimensions[get_column_letter(ci)].width = 14
        else:
            ws2.column_dimensions[get_column_letter(ci)].width = 18
    ws2.row_dimensions[2].height = 18
    ws2.freeze_panes = "A3"

    prop_ids = [str(p["id"]) for p in proposals]

    for ri, item in enumerate(items, start=3):
        bg = LINHAS[ri % 2]
        precos = item.get("precos", {})
        totais = item.get("totais", {})
        price_vals = [v for v in precos.values() if v is not None]

        _td(ws2, ri, 1, item["numero_item"], bg, "@", "center")
        _td(ws2, ri, 2, item.get("descricao", "")[:10], bg)
        _td(ws2, ri, 3, item.get("descricao", ""), bg)
        _td(ws2, ri, 4, item.get("unidade", ""), bg, "@", "center")
        _td(ws2, ri, 5, item.get("quantidade"), bg, NUM, "right")
        _td(ws2, ri, 6, item.get("preco_referencia"), bg, BRL, "right")
        _td(ws2, ri, 7, item.get("valor_referencia"), bg, BRL, "right")

        col = len(fixed_cols) + 1
        for pid in prop_ids:
            pu = precos.get(pid)
            tot = totais.get(pid)
            c_pu = _td(ws2, ri, col, pu, bg, BRL if pu else "@", "right")
            _td(ws2, ri, col+1, tot, bg, BRL if tot else "@", "right")
            if price_vals and pu is not None:
                if pu == min(price_vals):
                    c_pu.font = _font("15803D", bold=True, size=9)
                elif pu == max(price_vals):
                    c_pu.font = _font("DC2626", bold=False, size=9)
            col += 2

        _td(ws2, ri, col,   item.get("preco_minimo"), bg, BRL, "right")
        _td(ws2, ri, col+1, item.get("preco_maximo"), bg, BRL, "right")
        _td(ws2, ri, col+2, item.get("preco_medio"),  bg, BRL, "right")
        _td(ws2, ri, col+3, item.get("desvio_padrao"), bg, BRL, "right")
        _td(ws2, ri, col+4, item.get("categoria", ""), bg)
        _td(ws2, ri, col+5, item.get("disciplina", ""), bg)

    # ── ABA 3: Cherry Picking ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Cherry Picking")
    ws3.sheet_view.showGridLines = False

    cp_headers = ["Item", "Descrição", "Un", "Qtd", "Melhor P.Unit. COM REIDI (R$)",
                  "Fornecedor (menor preço)", "Total Cherry Pick (R$)",
                  "Ref. Total (R$)", "Economia vs Ref. (R$)", "Desvio vs Ref. (%)"]
    _title_row(ws3, len(cp_headers), f"CHERRY PICKING  ·  {project_name.upper()}")
    for ci, h in enumerate(cp_headers, 1):
        _th(ws3, 2, ci, h, bg="7C3AED")
        ws3.column_dimensions[get_column_letter(ci)].width = [8,48,6,10,22,30,20,16,18,16][ci-1]
    ws3.row_dimensions[2].height = 18
    ws3.freeze_panes = "A3"

    prop_name_map = {str(p["id"]): p["empresa"] for p in proposals}
    cp_run_total = 0.0

    for ri, item in enumerate(items, start=3):
        bg = LINHAS[ri % 2]
        precos = {k: v for k, v in item.get("precos", {}).items() if v is not None}
        if not precos:
            _td(ws3, ri, 1, item["numero_item"], bg, "@", "center")
            _td(ws3, ri, 2, item.get("descricao", ""), bg)
            _td(ws3, ri, 3, item.get("unidade", ""), bg, "@", "center")
            _td(ws3, ri, 4, item.get("quantidade"), bg, NUM, "right")
            for col in range(5, 11):
                _td(ws3, ri, col, "Sem cotação", bg, "@", "center")
            continue
        min_pid = min(precos, key=lambda k: precos[k])
        min_pu = precos[min_pid]
        qty = float(item.get("quantidade", 0))
        cp_line = qty * float(min_pu)
        ref_total = float(item.get("valor_referencia", 0) or 0)
        eco = ref_total - cp_line if ref_total else None
        eco_pct = (eco / ref_total * 100) if (eco is not None and ref_total) else None
        cp_run_total += cp_line

        _td(ws3, ri, 1, item["numero_item"], bg, "@", "center")
        _td(ws3, ri, 2, item.get("descricao", ""), bg)
        _td(ws3, ri, 3, item.get("unidade", ""), bg, "@", "center")
        _td(ws3, ri, 4, qty, bg, NUM, "right")
        _td(ws3, ri, 5, min_pu, bg, BRL, "right")
        _td(ws3, ri, 6, prop_name_map.get(min_pid, min_pid), bg)
        _td(ws3, ri, 7, cp_line, bg, BRL, "right")
        _td(ws3, ri, 8, ref_total if ref_total else None, bg, BRL, "right")
        c_eco = _td(ws3, ri, 9, eco, bg, BRL, "right")
        c_pct = _td(ws3, ri, 10, eco_pct, bg, PCT, "right")
        if eco and eco > 0:
            c_eco.font = _font("15803D", bold=True, size=9)
            c_pct.font = _font("15803D", bold=True, size=9)

    tr = len(items) + 3
    ws3.merge_cells(f"A{tr}:F{tr}")
    c = ws3[f"A{tr}"]
    c.value = "TOTAL CHERRY PICKING"
    c.fill = _fill(AZUL_TITULO); c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("right")
    t = ws3.cell(row=tr, column=7, value=cp_run_total)
    t.fill = _fill(AZUL_TITULO); t.font = _font("FFFFFF", bold=True, size=10)
    t.number_format = BRL; t.alignment = _align("right")

    # ── ABA 4: Por Fornecedor ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Por Fornecedor")
    ws4.sheet_view.showGridLines = False

    forn_headers = ["Posição", "Empresa", "Total COM REIDI (R$)", "Itens Cotados",
                    "Itens Sem Cotação", "Abaixo da Ref.", "Acima da Ref.",
                    "Desvio Médio vs Ref. (%)", "BDI Global (%)"]
    _title_row(ws4, len(forn_headers), f"ANÁLISE POR FORNECEDOR  ·  {project_name.upper()}")
    for ci, h in enumerate(forn_headers, 1):
        _th(ws4, 2, ci, h, bg="0F766E")
        ws4.column_dimensions[get_column_letter(ci)].width = [8,32,20,14,14,14,14,22,14][ci-1]
    ws4.row_dimensions[2].height = 18
    ws4.freeze_panes = "A3"

    sorted_forn = sorted(proposals, key=lambda p: p.get("valor_total", 0))
    n_items = len(items)

    for ri, p in enumerate(sorted_forn, start=3):
        bg = LINHAS[ri % 2]
        pid = str(p["id"])
        priced = sum(1 for item in items if item.get("precos", {}).get(pid) is not None)
        missing = n_items - priced
        below = 0; above = 0; devs = []
        for item in items:
            pu = item.get("precos", {}).get(pid)
            ref = item.get("preco_referencia")
            if pu is not None and ref and ref > 0:
                d = (float(pu) - float(ref)) / float(ref) * 100
                devs.append(d)
                if d < 0: below += 1
                else: above += 1
        avg_dev = sum(devs) / len(devs) if devs else None

        _td(ws4, ri, 1, ri-2, bg, "@", "center", bold=(ri==3))
        _td(ws4, ri, 2, p["empresa"], bg)
        _td(ws4, ri, 3, p.get("valor_total", 0), bg, BRL, "right", bold=(ri==3))
        _td(ws4, ri, 4, priced, bg, "@", "center")
        _td(ws4, ri, 5, missing, bg, "@", "center")
        _td(ws4, ri, 6, below, bg, "@", "center")
        _td(ws4, ri, 7, above, bg, "@", "center")
        c_dev = _td(ws4, ri, 8, avg_dev, bg, PCT, "right")
        if avg_dev is not None:
            c_dev.font = _font("15803D" if avg_dev < 0 else "DC2626", size=9)
        _td(ws4, ri, 9, p.get("bdi_global", 0), bg, PCT, "right")

    # ── ABA 5: Curva ABC ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Curva ABC")
    ws5.sheet_view.showGridLines = False

    abc_items = pareto_data.get("items", [])
    abc_headers = ["Pos.", "Item", "Descrição", "Un", "Qtd", "Preço Médio COM REIDI (R$)",
                   "Valor Total (R$)", "% Item", "% Acumulado", "Classe", "Categoria", "Disciplina"]
    _title_row(ws5, len(abc_headers), f"CURVA ABC  ·  {project_name.upper()}")
    for ci, h in enumerate(abc_headers, 1):
        _th(ws5, 2, ci, h)
        ws5.column_dimensions[get_column_letter(ci)].width = [6,8,46,6,10,22,18,10,12,8,20,16][ci-1]
    ws5.row_dimensions[2].height = 18
    ws5.freeze_panes = "A3"

    ABC_COLORS = {"A": ("DCFCE7", "166534"), "B": ("DBEAFE", "1E40AF"), "C": ("F3F4F6", "374151")}

    for ri, item in enumerate(abc_items, start=3):
        cls = item.get("classe", "C")
        bg_cls, fc_cls = ABC_COLORS.get(cls, ("FFFFFF", "000000"))
        bg = LINHAS[ri % 2]
        _td(ws5, ri, 1,  item.get("posicao"), bg, "@", "center")
        _td(ws5, ri, 2,  item.get("numero_item", ""), bg, "@", "center")
        _td(ws5, ri, 3,  item.get("descricao", ""), bg)
        _td(ws5, ri, 4,  item.get("unidade", ""), bg, "@", "center")
        _td(ws5, ri, 5,  item.get("quantidade"), bg, NUM, "right")
        _td(ws5, ri, 6,  item.get("preco_medio"), bg, BRL, "right")
        _td(ws5, ri, 7,  item.get("valor_total"), bg, BRL, "right")
        _td(ws5, ri, 8,  item.get("percentual"), bg, PCT, "right")
        _td(ws5, ri, 9,  item.get("percentual_acumulado"), bg, PCT, "right")
        c_cls = ws5.cell(row=ri, column=10, value=cls)
        c_cls.fill = _fill(bg_cls); c_cls.font = _font(fc_cls, bold=True, size=9)
        c_cls.alignment = _align("center"); c_cls.border = _border()
        _td(ws5, ri, 11, item.get("categoria", ""), bg)
        _td(ws5, ri, 12, item.get("disciplina", ""), bg)

    return _save(wb)


# ── Utilitários internos ──────────────────────────────────────────────────────

def gerar_baseline_excel(entries: list) -> io.BytesIO:
    """
    Gera o Excel do Baseline de Contratos (multi-aba).
    Preços exibidos são COM REIDI (já normalizados no router analytics.py).
    """
    from collections import defaultdict
    import datetime as dt

    wb = Workbook()
    wb.remove(wb.active)

    GOLD   = "B45309"
    GOLD_L = "FEF3C7"
    GOLD_H = "FDE68A"
    DARK   = "1E293B"
    TIPO_LABELS = {
        "INFRAESTRUTURA": "Infraestrutura",
        "EDIFICACAO": "Edificação",
        "OBRA_DE_ARTE": "Obra de Arte Especial",
    }

    def _hdr(ws, row, cols, bg, fg="FFFFFF", bold=True, size=9):
        for ci, label in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=label)
            c.fill = _fill(bg)
            c.font = _font(fg, bold=bold, size=size)
            c.alignment = _align("center")
            c.border = _border()
        ws.row_dimensions[row].height = 18

    def _row(ws, row, vals, bg, fmts=None):
        for ci, val in enumerate(vals, 1):
            fmt = (fmts or {}).get(ci, "@")
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = _fill(bg)
            c.font = _font(size=9)
            c.border = _border()
            c.number_format = FMT.get(fmt, fmt)
            c.alignment = _align("right" if fmt not in ("@", "text") else "left")

    ws1 = wb.create_sheet("Contratos")
    ws1.sheet_view.showGridLines = False

    title_cols = ["Projeto", "Nº TR", "Tipo de Obra", "Proponente", "CNPJ",
                  "BDI (%)", "Valor Total COM REIDI (R$)", "Extensão (km)", "R$/km", "Data Premiação"]
    ncols = len(title_cols)
    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws1["A1"]
    c.value = "BASELINE — HISTÓRICO DE CONTRATOS PREMIADOS"
    c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center"); ws1.row_dimensions[1].height = 28

    _hdr(ws1, 2, title_cols, GOLD)
    widths = [40, 18, 20, 35, 20, 10, 22, 14, 14, 18]
    for ci, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A3"

    total_geral = 0.0
    for ri, e in enumerate(entries, start=3):
        bg = GOLD_L if ri % 2 == 0 else GOLD_H
        ext = e.get("extensao_km")
        vt  = e.get("valor_total", 0)
        total_geral += vt
        rk = (vt / ext) if ext else None
        ts = e.get("data_premiacao", "")
        try:
            ts = dt.datetime.fromisoformat(ts).strftime("%d/%m/%Y")
        except Exception:
            pass
        _row(ws1, ri,
             [e["project_nome"], e.get("numero_licitacao") or "—",
              TIPO_LABELS.get(e["tipo_obra"], e["tipo_obra"]),
              e["empresa"], e.get("cnpj") or "—",
              e["bdi_global"], vt, ext or "—", rk, ts],
             bg,
             {6: "pct", 7: "brl2", 9: "brl2"})

    tr = len(entries) + 3
    ws1.merge_cells(f"A{tr}:F{tr}")
    c = ws1[f"A{tr}"]
    c.value = "TOTAL GERAL"
    c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=10)
    c.alignment = _align("right")
    ct = ws1.cell(row=tr, column=7, value=total_geral)
    ct.fill = _fill(DARK); ct.font = _font("FFFFFF", bold=True, size=10)
    ct.number_format = FMT["brl2"]

    ws2 = wb.create_sheet("Itens Detalhados")
    ws2.sheet_view.showGridLines = False

    item_cols = ["Projeto", "Proponente", "Item", "Descrição", "Un",
                 "Qtd", "Categoria", "Disciplina", "P.Unit. COM REIDI (R$)", "Total COM REIDI (R$)"]
    ws2.merge_cells(f"A1:{get_column_letter(len(item_cols))}1")
    c = ws2["A1"]
    c.value = "BASELINE — ITENS DETALHADOS"
    c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center"); ws2.row_dimensions[1].height = 28

    _hdr(ws2, 2, item_cols, AZUL_HEADER)
    iwidths = [35, 28, 8, 52, 7, 12, 20, 16, 20, 20]
    for ci, w in enumerate(iwidths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A3"

    ri2 = 3
    for e in entries:
        for item in e.get("items", []):
            bg = AZUL_LINHA_A if ri2 % 2 == 0 else AZUL_LINHA_B
            _row(ws2, ri2,
                 [e["project_nome"], e["empresa"], item["numero_item"],
                  item["descricao"], item["unidade"], item["quantidade"],
                  item.get("categoria") or "—", item.get("disciplina") or "—",
                  item.get("preco_unitario"), item["preco_total"]],
                 bg, {6: "num4", 9: "brl4", 10: "brl2"})
            ri2 += 1

    ws3 = wb.create_sheet("Por Disciplina")
    ws3.sheet_view.showGridLines = False

    disc_agg = defaultdict(float)
    for e in entries:
        for item in e.get("items", []):
            disc = item.get("disciplina") or "Sem disciplina"
            disc_agg[disc] += item.get("preco_total", 0)

    total_disc = sum(disc_agg.values()) or 1
    d_cols = ["Disciplina", "Valor Total COM REIDI (R$)", "Participação (%)"]
    ws3.merge_cells("A1:C1")
    c = ws3["A1"]
    c.value = "BASELINE — DISTRIBUIÇÃO POR DISCIPLINA"
    c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center"); ws3.row_dimensions[1].height = 28
    _hdr(ws3, 2, d_cols, AZUL_HEADER)
    for ci, w in enumerate([30, 24, 16], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    for ri3, (disc, val) in enumerate(sorted(disc_agg.items(), key=lambda x: -x[1]), start=3):
        bg = AZUL_LINHA_A if ri3 % 2 == 0 else AZUL_LINHA_B
        pct = val / total_disc * 100
        _row(ws3, ri3, [disc, val, pct], bg, {2: "brl2", 3: "pct"})

    ws4 = wb.create_sheet("Por Categoria")
    ws4.sheet_view.showGridLines = False

    cat_agg = defaultdict(float)
    for e in entries:
        for item in e.get("items", []):
            cat = item.get("categoria") or "Sem categoria"
            cat_agg[cat] += item.get("preco_total", 0)

    total_cat = sum(cat_agg.values()) or 1
    c_cols = ["Categoria", "Valor Total COM REIDI (R$)", "Participação (%)"]
    ws4.merge_cells("A1:C1")
    c = ws4["A1"]
    c.value = "BASELINE — DISTRIBUIÇÃO POR CATEGORIA"
    c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=13)
    c.alignment = _align("center"); ws4.row_dimensions[1].height = 28
    _hdr(ws4, 2, c_cols, VERDE_HEADER)
    for ci, w in enumerate([30, 24, 16], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    for ri4, (cat, val) in enumerate(sorted(cat_agg.items(), key=lambda x: -x[1]), start=3):
        bg = VERDE_LINHA_A if ri4 % 2 == 0 else VERDE_LINHA_B
        pct = val / total_cat * 100
        _row(ws4, ri4, [cat, val, pct], bg, {2: "brl2", 3: "pct"})

    km_entries = [e for e in entries if e.get("extensao_km")]
    if km_entries:
        ws5 = wb.create_sheet("Custo por km")
        ws5.sheet_view.showGridLines = False

        km_cols = ["Projeto", "Tipo", "Extensão (km)", "Valor COM REIDI (R$)", "R$/km",
                   "Disciplina", "Valor Disciplina (R$)", "R$/km Disciplina"]
        ws5.merge_cells(f"A1:{get_column_letter(len(km_cols))}1")
        c = ws5["A1"]
        c.value = "BASELINE — CUSTO POR QUILÔMETRO"
        c.fill = _fill(DARK); c.font = _font("FFFFFF", bold=True, size=13)
        c.alignment = _align("center"); ws5.row_dimensions[1].height = 28
        _hdr(ws5, 2, km_cols, GOLD)
        for ci, w in enumerate([35, 20, 14, 22, 16, 20, 20, 16], 1):
            ws5.column_dimensions[get_column_letter(ci)].width = w
        ws5.freeze_panes = "A3"

        ri5 = 3
        for e in km_entries:
            ext = e["extensao_km"]
            vt  = e["valor_total"]
            rk  = vt / ext
            disc_vals = defaultdict(float)
            for item in e.get("items", []):
                d = item.get("disciplina") or "Sem disciplina"
                disc_vals[d] += item.get("preco_total", 0)

            first = True
            for disc, dval in sorted(disc_vals.items(), key=lambda x: -x[1]):
                bg = GOLD_L if ri5 % 2 == 0 else GOLD_H
                _row(ws5, ri5,
                     [e["project_nome"] if first else "",
                      TIPO_LABELS.get(e["tipo_obra"], e["tipo_obra"]) if first else "",
                      ext if first else "", vt if first else "", rk if first else "",
                      disc, dval, dval / ext],
                     bg,
                     {3: "num4", 4: "brl2", 5: "brl2", 7: "brl2", 8: "brl2"})
                first = False
                ri5 += 1

    return _save(wb)


def _save(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _find_header(ws, mapa: dict, required: set):
    """Localiza a linha de cabeçalho e retorna (row_num, {col_idx: field_name})."""
    for row in ws.iter_rows(min_row=1, max_row=12):
        col_map = {}
        for cell in row:
            key = str(cell.value or "").strip().lower()
            field = mapa.get(key)
            if field:
                col_map[cell.column] = field
        if required.issubset(col_map.values()):
            return cell.row, col_map
    raise ValueError(
        f"Cabeçalhos obrigatórios não encontrados: {required}. "
        "Verifique se o arquivo usa o modelo padrão."
    )


def _str(v) -> Optional[str]:
    s = str(v or "").strip()
    return s if s else None


def _float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").replace("R$", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None
